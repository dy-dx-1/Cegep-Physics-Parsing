from math import expm1
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import load_workbook

# Déclaration des constantes 
T = 293.15                   # Température en kelvins  
Kb = 1.38064852 * (10**-23)  # Constante de Boltzmann 
q = 1.60217662 * (10**-19)   # Charge élémentaire 
Phi = float(q/(Kb * T))      # Constante Phi 

# Enregistrement des variables nécéssaires 
Vco = float(input("Valeur de V circuit ouvert en Volts: "))    
Isc = float(input("Valeur de Isc en mA: "))                              # Enregistrement de Vco et Isc 
MINC = int(input("Numéro représentant case quadrant 2 en 'x': "))
MAXC = int(input("Numéro représentant case quadrant 1 en 'x': "))
MINR = int(input("Numéro représentant case quadrant 2 en 'y': "))
MAXR = int(input("Numéro représentant case quadrant 3 en 'y': "))

#calcul de Io et conversion en notation scientifique 
Io = float(Isc/expm1(Vco * Phi))
Io_SCT_Notation = '{:.2e}'.format(Io)   
# Deux Io, un pour "print" et un autre pour les calculs 

# Enregistrement de l'équation dans la variable "text"
text = "Équation: I = " + str(Isc) + " - "+ Io_SCT_Notation + " * (exp("+str(Phi)[:7] + " * v) - 1 )"

# Équation théorique de la courbe IV 
IV = lambda v: Isc - ( Io * (expm1(Phi * v)))

# Création des listes utilisées afin de faire la courbe théorique 
Valeurs_X = []
Valeurs_Y = []
for u in np.arange(0, Vco+1, Vco/1000):  
    try: 
        Valeurs_Y.append(IV(u))
        Valeurs_X.append(u)
    except OverflowError:
        # Si le nombre est trop grand pour être calculé - ne pas calculer le point (quelques pts manquants avec 1000 bonds ne seront pas significatifs)
        pass


# Préparation du graphique 

plt.axis([0, Vco+1, 0, Isc+1])   # Vco+1 pour laisser de l'espace à droite de abcisse à l'origine 
plt.grid(True)
plt.plot(Valeurs_X, Valeurs_Y, 'r-')       
plt.xlabel("Différence de potentiel aux bornes (V)")
plt.ylabel("Intensité de courant (mA)")                 # Noms des axes 
title = input("Titre du graphique: ")
plt.title(title)                                 


# Calcul de l'aire optimale
list_aire = []
for n in range(len(Valeurs_X)):
    Aire = Valeurs_X[n] * Valeurs_Y[n] 
    list_aire.append(Aire)
# Maintenant, trouvons l'aire la plus grande et retrouvons le n appropié 
indice_max = list_aire.index(max(list_aire))
# Imprimer la valeur de x correspondant à l'aire maximale 
print(f"V optimal: {Valeurs_X[indice_max]}")

# Prenons maintenant les données expérimentales 
excel = load_workbook(filename = r"C:\Users\Nicolas\Desktop\Phy_EXCEL.xlsx", data_only= True)               # location du fichier excel 
sheet = excel.active # Il faut fermer le excel dans la page appropiée avant de commencer les calculs 

#Récupération des données et enregistrement dans des listes 
Valeurs_X_E = []
Valeurs_Y_E = []
Incert_X = []
Incert_Y = []
c = 1 
for row in sheet.iter_cols(min_row=MINR, max_row=MAXR, min_col=MINC, max_col=MAXC, values_only= True):
    if c==1: 
        for val in row:
            Valeurs_X_E.append(val)
    elif c==2:
        for val in row:
            Incert_X.append(val)
    elif c==3:
        for val in row:
            Valeurs_Y_E.append(val)
    elif c==4: 
        for val in row:
            Incert_Y.append(val)
    c+=1 

# Courbe expérimentale et barres d'erreur
plt.errorbar(Valeurs_X_E, Valeurs_Y_E, yerr=Incert_Y, xerr=Incert_X) 
plt.plot(Valeurs_X_E, Valeurs_Y_E, 'b-')   

# légendes 
red_patch = mpatches.Patch(color='red', label='Courbe théorique')
blue_patch = mpatches.Patch(color = 'blue', label = 'Courbe expérimentale')
first_legend = plt.legend(handles=[red_patch], loc = 1)
ax = plt.gca().add_artist(first_legend)
plt.legend(handles=[blue_patch], loc = 2)

print(text.replace("e", "*10^", 1)) # Imprimer l'équation avec un format facile à recopier dans excel ou geogebra 

plt.show()  # Montrer le graphique 